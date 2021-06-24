# import libs
from sqlalchemy import create_engine, text
import pyperclip as pc
import psycopg2
import pandas as pd
import datetime
import os
import glob
os.path.abspath(os.getcwd())
import sys
import errno
from pathlib import Path
from flask import make_response,jsonify
import sandesh



# Crea la clase
class queries:
    # Crea una constante que contiene el texto por defecto para el mensaje
    BLOCK = {
        "type": "section",
        "text": {
            "type": "mrkdwn",
            "text": (
                "_Procesando..._\n\n"
            ),
        },
    }

    # Constructor
    def __init__(self,channel,**kwargs):
        self.channel = channel
        self.fuero = kwargs['fuero']
        

    def _get_query(self,**kwargs):
        fuero = kwargs['fuero']
        keyword = kwargs['keyword']

        #ddedbs
        hostname = os.getenv("ddedbs_ip")
        username = os.getenv("ddedbs_username")
        password = os.getenv("ddedbs_pass")

        #vero
        hostname_vero = os.getenv("pc_vl_ip")
        username_vero = os.getenv("pc_vl_username")
        password_vero = os.getenv("pc_vl_pass")

        ##### vuelve al directorio raiz si es que esta en el directorio output ####
        cwd = os.getcwd()
        if 'output' in cwd:
            os.chdir("..")

        #### toma el archivo sql correspondiente ####
        query_path = os.path.join(os.getcwd(),'queries')
        if keyword:
            keywordsql = keyword + '.sql'
            sql_path = os.path.join(query_path,keywordsql)
            if keyword[len(keyword)-1] == '_': #si el keyword termina en _ no tiene parametros
                parameters = False
            else:
                parameters = True
                

        #### BD connections ####
        if '_' in fuero: # si esta _ en fuero es porque se esta consultando los backups, la conexion va a la pc que guarda los backups
            con = psycopg2.connect(host=hostname_vero, user=username_vero, password=password_vero)
            lista_db = 'SELECT datname FROM pg_database WHERE datistemplate = false'
            bases = pd.read_sql(lista_db,con)
            dbs = [base for base in bases.datname if base.startswith('sae')]
            print(f'ESTA ES LA BASE DE CONSULTA {dbs}')
            hostname = hostname_vero
            username = username_vero
            password = password_vero
        else: # si _ no esta en fuero entonces se consulta las bases ddedbs
            print(hostname,username,password)
            con = psycopg2.connect(host=hostname, user=username, password=password, dbname='saeciv')
            lista_db = '''SELECT datname FROM pg_database WHERE datistemplate = false '''
            bases = pd.read_sql(lista_db,con)
            dbs = [base for base in bases.datname if (base.startswith('sae') or base.startswith('oga')) and not base.startswith('saemed') and base not in ('saejes','saepjt')]
        dfs = pd.DataFrame()


        #### corre la consulta ####
        for  fila in dbs:
            if fuero == fila:
                print(f'Iniciando lectura de base de datos {fila}', end='\n')
                msg = ''
                msg = f'Iniciando lectura de base de datos {fila}'
                #sandesh.send(msg, webhook = "https://hooks.slack.com/services/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")
                print(f'\n datos conexion: {hostname} {username} {password} {fila}')
                con_b = psycopg2.connect(host=hostname, user=username, password=password, dbname = fila)
                query_file = open(sql_path, 'r', encoding="latin-1")
                texto = query_file.read()
                
                if parameters == True:
                    fechadesde = kwargs['fechadesde'] 
                    fechahasta = kwargs['fechahasta']
                    texto = texto.format(fechadesde,fechahasta)
                else:
                    texto = texto
                
                df = pd.read_sql(texto,con_b)
                print(f'DF: {df}')
                dfs = dfs.append(df)
                con_b.close()
        #### guarda el resultado de la consulta en un excel y agrega la query en un nuevo sheet (metadata) ####
       
        fechadesde = kwargs['fechadesde'] 
        fechahasta = kwargs['fechahasta']
        if parameters == True:
            nombre = f'{keyword}_{fuero}_desde_{fechadesde}_hasta_{fechahasta}.xlsx'
        else:
            nombre = f'{keyword}_{fuero}.xlsx'
        
        cwd = os.getcwd()
        if 'output' not in cwd: # se posiciona en el directorio output para guardar el archivo excel
            os.chdir("output")
        msg = dfs.to_excel(nombre, sheet_name = 'data', index=False, header= True)
        from openpyxl import load_workbook
        wb2 = load_workbook(nombre)
        ws=wb2.create_sheet('metadata')
        ws.cell(row=1, column=1).value = 'query'
        ws.cell(row=2, column=1).value = texto
        wb2.save(nombre)
        msg = wb2
    
        if 'output' in os.getcwd():
            os.chdir("..") # vuelve al directorio raiz del proyecto 
        
        return {"type": "section", "text": {"type": "mrkdwn", "file": msg}}

    #### Crea y devuelve el payload como un diccionario. ####
    def get_message_payload(self,**kwargs):
        return {
            "X-Slack-No-Retry": 1, 
            "channel": self.channel,
            "blocks": [
                self.BLOCK,
                *self._get_query(**kwargs),
            ],
        }